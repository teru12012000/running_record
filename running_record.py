import tkinter as tk
import datetime
import openpyxl
import os
import calendar
import tkinter.messagebox as messa

File_name="practice.xlsx"
#日付取得
dt_time=datetime.date.today()
text_dt=f'{dt_time.year}/{dt_time.month}/{dt_time.day}'
sheet_name=f'{dt_time.year}年{dt_time.month}月'
today=(int)(dt_time.day)
last_day=calendar.monthrange(dt_time.year,dt_time.month)[1]
menu=[
  sheet_name,
  'メニュー1',
  'メニュー2',
  'メニュー3',
  '合計距離(km)',
]

def onchange(text,num):
  global today,ws,wb
  txt=text.get()
  c=ws.cell(row=today+1,column=num+1)
  c.value=txt
  messa.showinfo("報告","登録しました。")
  wb.save(File_name)
  
def ondelete(text,num):
    global today,ws,wb
    c=ws.cell(row=today+1,column=num+1)
    c.value=None
    messa.showinfo("報告","削除しました。")
    wb.save(File_name)
    
def onchangetext(text,num):
  global today,ws,wb
  txt=text.get('1.0', 'end')
  c=ws.cell(row=today+1,column=num+1)
  messa.showinfo("報告","登録しました。")
  c.value=txt
  wb.save(File_name)
  
def ondeletetext(text,num):
    global today,ws,wb
    c=ws.cell(row=today+1,column=num+1)
    messa.showinfo("報告","削除しました、")
    c.value=None
    wb.save(File_name)
    
def onnumber(text):
  global today,ws,wb,last_day
  txt=text.get()
  c=ws.cell(row=today+1,column=5)
  c.value=int(txt)
  day=str(last_day+1)
  val=f'=sum(E2:E{day})'
  
  c=ws.cell(row=last_day+2,column=4)
  c.value='今月合計(km)'
  c=ws.cell(row=last_day+2,column=5)
  c.value=f'{val}'
  messa.showinfo("報告","登録しました。")
  wb.save(File_name)

def ondeletenumber():
  global today,ws,wb,last_day
  
  c=ws.cell(row=today+1,column=5)
  c.value=None
  day=str(last_day+1)
  val=f'=sum(E2:E{day})'
  
  c=ws.cell(row=last_day+2,column=4)
  c.value='今月合計'
  c=ws.cell(row=last_day+2,column=5)
  c.value=f'{val}km'
  messa.showinfo("報告","削除しました。")
  wb.save(File_name)  
#基本設定
app=tk.Tk()
app.geometry("1000x800")
app.title("ランニングの記録")
app.resizable(False,False)



#Excelの準備
if os.path.exists(File_name):
  wb=openpyxl.load_workbook(File_name)
else:
  wb=openpyxl.Workbook()
  ws=wb.active
  ws.title=sheet_name

wses=wb.sheetnames

if sheet_name in wses:
  ws=wb[sheet_name]
else:
  wb.create_chartsheet(title=sheet_name)
  ws=wb[sheet_name]
#wb.save(File_name)
#Excelの基本的なメニューの挿入
for i in range(len(menu)):
  c1=ws.cell(row=1,column=i+1)
  c1.value=menu[i]
for i in range(last_day+1):
  ws.merge_cells(start_row=i+1, end_row=i+1, start_column=6, end_column=14)
thought=ws["F1"]
thought.value='感想'
for i in range(last_day):
  c1=ws.cell(row=i+2,column=1)
  c1.value=str(i+1)
#日付の表示
paint_date=tk.Label(app,text=text_dt)
paint_date.place(
  x=900,
  y=10,
)
message=tk.Label(app,text="本日の練習の記録(別日はexcel編集をすること)")
#メニュー1の入力
menu1_label=tk.Label(app,text='1つ目のメニュー:')
menu1_entry=tk.Entry(app)
menu1_change=tk.Button(app,text='登録' ,command=lambda:onchange(menu1_entry,1))
menu1_delete=tk.Button(app,text='記録を削除',command=lambda:ondelete(menu1_entry,1))
menu1_label.pack()
menu1_entry.pack()
menu1_change.pack()
menu1_delete.pack()

#メニュー2の入力
menu2_label=tk.Label(app,text='2つ目のメニュー:')
menu2_entry=tk.Entry(app)
menu2_change=tk.Button(app,text='登録' ,command=lambda:onchange(menu2_entry,2))
menu2_delete=tk.Button(app,text='記録を削除',command=lambda:ondelete(menu2_entry,2))
menu2_label.pack()
menu2_entry.pack()
menu2_change.pack()
menu2_delete.pack()

#メニュー3の入力
menu3_label=tk.Label(app,text='3つ目のメニュー:')
menu3_entry=tk.Entry(app)
menu3_change=tk.Button(app,text='登録' ,command=lambda:onchange(menu3_entry,3))
menu3_delete=tk.Button(app,text='記録を削除',command=lambda:ondelete(menu3_entry,3))
menu3_label.pack()
menu3_entry.pack()
menu3_change.pack()
menu3_delete.pack()

#距離の入力
distance_label=tk.Label(app,text='合計走行距離(km)※数字のみ')
distance_entry=tk.Entry(app)
distance_change=tk.Button(app,text='登録', command=lambda:onnumber(distance_entry))
distance_delete=tk.Button(app,text='記録を削除', command=lambda:ondeletenumber())
distance_label.pack()
distance_entry.pack()
distance_change.pack()
distance_delete.pack()


#感想の入力
thought_label=tk.Label(app,text='感想')
thought_entry=tk.Text(app)
thought_change=tk.Button(app,text='登録',command=lambda:onchangetext(thought_entry,5))
thought_delete=tk.Button(app,text='記録を削除',command=lambda:ondeletetext(thought_entry,5))
thought_label.pack()
thought_entry.pack()
thought_change.pack()
thought_delete.pack()



wb.save(File_name)
app.mainloop()